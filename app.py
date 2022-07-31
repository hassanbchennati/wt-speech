import speech_recognition as sr
from flask import logging, Flask, render_template, request, flash
import pocketsphinx


import numpy as np
import math

import formulas
import difflib

import openpyxl

app = Flask(__name__)
app.secret_key = "VatsalParsaniya"

keywords = [
    ("plant", 1), 
    ("one", 0), 
]
def AramcoSheet12(csvpath):
    return("TESTING THE FUCTION")



def AramcoSheet1 (csvpath,d): 
        
        '''
    Read Aramco cvs data: How much NGL we can produce in UGP if Khuff diverted to LP Letdown?  
    and take all function needed for the calulation to process it with the user input
    
    Attributes:
    -----------
        csvpath: path to the csv file 
    '''
        
        userInputs = {"Enter Associated Gas Feed Rate" :"G32", "Enter Khuff to LP Letdown": "G38" }
        
        readingInputs= {}
        
        userOutput = {"TOTAL feed rate to LRUs": "U31", "Intercepted %C2 Recovery": "U37",
                      "NGL Production ": "U43" ,"Additional NGL following Khuff to LP Letdown": "U50", 
                     "Value of the additional NGL" : "U56"}
        
        calculationOrder = ["U31", "U37", "AE20", "U43", "U50", "U56"]
        calculationOutput = {}
        
        finalResult = {}
    
        # read the excel sheet 
        wb = openpyxl.load_workbook(csvpath)
        ws = wb['Input']
    
        
        readingInputs['G32'] = d[0]
        readingInputs['G38']=  d[1]
        #print(readingInputs)
        
        for i in calculationOrder: 
            inArgs = {}
            func = formulas.Parser().ast(ws[i].value)[1].compile()
            funcArgs = list(func.inputs)
          #  print("func input", list(func.inputs))

            
            for cell in readingInputs: 
                if cell in funcArgs: 
                    inArgs[cell]=readingInputs[cell] # make sure 
                   # print("done with reading input")
                    
            
            for j in funcArgs: 
                if j == "Q5": 
                    inArgs["Q5"]=0
                
                elif j in calculationOutput: 
                    inArgs[j] = calculationOutput[j]

            
            if len(inArgs)== len(funcArgs):
                # make sure the args in the correct order 
                reordered_args = {k: inArgs[k] for k in funcArgs}
                
                templist = []
                for k in reordered_args : 
                    templist.append(reordered_args[k])
                    
                
                calculationOutput[i] = func(*templist)
                
        
        for final in userOutput:
                finalResult[final] = str(calculationOutput[userOutput[final]])
        
        #print(finalResult)
        
        return(finalResult)
                




def AramcoSheet2(csvpath, digits): 
          
        '''
    Read Aramco cvs data: P&RMU NGL Spheres D-011 A/B Level Calcuations  
    and take all function needed for the calulation to process it with the user input
    
    Attributes:
    -----------
        csvpath: path to the csv file 
    '''
        
        
        historainInputCase = {"Initial Level (%)": "E8", "Final Level (%)": "E10"}
        historainOutput = {"final volume (bbl)": "E11", 
                          "Time (h)" : "E12", 
                          "Time (h)  [2 spheres]": "E13"}
        
    
        manualInputcase = {"Total Flow from LRU (MBD)":"E16",
                   "Total flow from Inlet (MBD)":"E17",
                  "Total NGL production (MBD)":"E18", 
                  "Initial Level (%)": "E20",
                  "Final Level (%)": "E22"}
        manualOutput = {"final volume (bbl)": "E23", 
                          "Time (h)" : "E24", 
                          "Time (h)  [2 spheres]": "E25"}
        
        readingInputs= []
        
#         HistorainValues = {"D55": 83.57, "D56": 82.93, "D57": 64.83, "D58" : 12.08, "D60": 241.70, 
#                           "D61" : 31.68, "D62" : 31.07}
        
        calculationOutput = {}
    
        # read the excel sheet 
        wb = openpyxl.load_workbook(csvpath , data_only = True)
        ws = wb['NGL Sphere']
        
          # take the input from the user 
    
        #userinput= input("Do you want to use the historain values for caculation? Y/N")
        userinput = "y"
        if userinput.lower() == "y":
            
            #read the needed input from the user 
            
            # for userinput in historainInputCase: 
            #     hisIn= input(userinput)
            #     readingInputs.append(float(hisIn)/100)
            
            # readingInputs.append(float(31.67564010620120)/100)
            # readingInputs.append(float(75)/100)
            readingInputs.append(float(digits[0])/100)
            readingInputs.append(float(digits[1])/100)


#             print(readingInputs)
            
            E8 =readingInputs[0]
            E10 =readingInputs[1]
            
            # assign the values from excel sheet 
                
            B5 = ws['B5'].value
            E5 = ws['E5'].value
            E7 = ws['E7'].value
            
            # initial volume (bbl)
            E9 = ( np.pi * ( math.pow( E8 * B5, 2) * ( 3 * B5 / 2 - E8 * B5 ) / 3) ) * (0.178108) # barrels

            # final volume (bbl)
            E11 = ( np.pi * ( math.pow( E10 * B5, 2) * ( 3 * B5 / 2 - E10 * B5 ) / 3) ) * (0.178108) # barrels

            # Time (h)  
            E12 = abs ( ( E11 - E9 ) / ( E7 * 1000 ) * 24 )

            # Time (h)  [2 spheres]
            E13 = E12*2
            
            # assign the results to the the output dictorary 
            
            for s in historainOutput: 
                if historainOutput[s] == 'E11': 
                    calculationOutput[s] = "{:.2f}".format(E11)
                elif historainOutput[s] == 'E12': 
                    calculationOutput[s] = "{:.2f}".format(E12)
                
                elif historainOutput[s] == 'E13': 
                    calculationOutput[s] = "{:.2f}".format(E13)
            
        elif userinput.lower() == "n":
  
            # read the needed input from the user
            count = 0 
            for userinput in manualInputcase: 
                mIn= input(userinput)
        
                # condition for Percentage cells 
                if count ==3 or count == 4 : 
                    readingInputs.append(float(mIn)/100)
                else: 
                    readingInputs.append(float(mIn))
                
                count += 1
                
            # assign the values from excel sheet     
            E16 = readingInputs[0]
            E17 = readingInputs[1]
            E18 = readingInputs[2]
            E19 =(E16+E17)-E18
            E20 = readingInputs[3]
            E22 = readingInputs[4]
            B5 = ws['B5'].value
            
            print("E16", E16, "E17", E17, "E18", E18, "E19", E19, "E20", E20, "E22", E22, "B5", B5)
            E21 = (  np.pi  * ( (E20 * B5 )**2 * ( 3 * B5 / 2 - E20 * B5 )/3 ) * (0.178108) ) # initial volume (bbl)
            E23 = ( np.pi * ( math.pow( E22 * B5, 2) * ( 3 * B5 / 2 - E22 * B5 ) / 3) ) * (0.178108) # barrels # final volume (bbl)
            E24 =abs ( ( E23 - E21 ) / ( E19 * 1000 ) * 24 ) # Time (h)  
            E25 = E24 * 2 # Time (h)  [2 spheres]
            
            
            for s in manualOutput: 
                if manualOutput[s] == 'E23': 
                    calculationOutput[s] = E23
                elif manualOutput[s] == 'E24': 
                    calculationOutput[s] = E24
                
                elif manualOutput[s] == 'E25': 
                    calculationOutput[s] = E25
        
            
            print(calculationOutput)
    
        return (calculationOutput)
        
def similarity(word, pattern):
    return difflib.SequenceMatcher(a=word.lower(), b=pattern.lower()).ratio()


def datacomparsion(lookup): 
    
    flag1 = "NGL LP Letdown feed rate LRUs Intercepted %C2 Recovery Production Additional Khuff Let down n"
    flag2 = "P&RMU RMU and P NGL pair spheres D 011 A B Level final volume time P M" 
    threshold = 0.6
    flag1count = 0 
    flag2count = 0 

    # extract the digits from string 

    nums = lookup.split()
    digits = []

    for i in nums: 
        if i.isdigit() == True : 
            digits.append(int(i))
        else: 
            pass 
    
    
    for word in flag1.split():
        for s in lookup.split(): 
            if similarity(word, s) > threshold:
                print(word)
                flag1count +=1 

    for word in flag2.split():
        for s in lookup.split(): 
            if similarity(word, s) > threshold:
                print(word)
                flag2count +=1 



    if flag1count> flag2count: 
        print("sheet 1 selected")
        result= AramcoSheet1("How much NGL we can produce in UGP.xlsm" , digits) # add the path
        print("result")
    elif flag1count< flag2count: 
        print("sheet 2 selected")
        result = AramcoSheet2("NGL Sphere Calculations (D-011 AB)_2.xlsx", digits)# add the path
    else: 
        print("your voice not clear")
        result = "your voice not clear"

    return (result)



@app.route('/')
def index():
    flash(" Welcome to Aramco's Robot Demo")
    return render_template('index.html')

@app.route('/audio_to_text/')
def audio_to_text():
    flash(" Press Start to start recording audio and press Stop to end recording audio")
    return render_template('audio_to_text.html')

@app.route('/audio', methods=['POST'])
def audio():
    r = sr.Recognizer()
    with open('upload/audio.wav', 'wb') as f:
        f.write(request.data)
  
    with sr.AudioFile('upload/audio.wav') as source:

        #r.adjust_for_ambient_noise(source)
        audio_data = r.record(source)
        text = r.recognize_google(audio_data, language='en-IN', show_all=True)

        print(text)
        #print("recognize_google" ,text["alternative"][-1]["transcript"])
        return_text = " Did you say : <br> "

        return_text = ""
        try:

            print("we are in try")
            #print("recognize_google" ,text["alternative"][-1]["transcript"])

            return_text = datacomparsion(text["alternative"][-2]["transcript"])
            
            print("this is the return text", return_text)


            # for num, texts in enumerate(text['alternative']):
            #     #return_text += str(num+1) +") " + texts['transcript']  + " <br> "
            #     if "level" in texts['transcript'] :
            #        test=AramcoSheet1("How much NGL we can produce in UGP.xlsm")
            #        text = 'TOTAL feed rate to LRUs is ' + str(test['TOTAL feed rate to LRUs']) + ","
            #        text = text + 'Intercepted C2 Recovery' + test['Intercepted %C2 Recovery'] + ","
            #        text = text + 'NGL Production ' + test['NGL Production '] + ","
            #        text = text + 'Additional NGL following Khuff to LP Letdown' + test['Additional NGL following Khuff to LP Letdown'] + ","
            #        text = text + '$ Value of the additional NGL' + test['$ Value of the additional NGL']

            #        return_text += text
            #        break
            #     if "flow" in texts['transcript'] :
            #        return_text += "The flow level is 1500"
            #        break

        except:
            return_text = " Sorry!!!! Voice not Detected "
        '''   
        #r.adjust_for_ambient_noise(source)
        audio_data = r.record(source)
        text = r.recognize_sphinx(audio_data, language='en-US', show_all=True)#,keyword_entries = keywords)
        print(text)
        return_text = " Did you say : <br> "
        return_text = ""
        for best, i in zip(text.nbest(), range(10)):
             print (best.hypstr, best.score)
             if "level" in best.hypstr:
             	return_text += "The plant level is 2000"
             	break	
        print('Best hypothesis segments: ', [(seg.word, seg.prob) for seg in text.seg()])
        '''
    
    print(str(return_text))    
    f1 = str(return_text).replace(',','<br/>')
    f1 = f1.replace('}','')
    f1 = f1.replace('{','')
    f1 = f1.replace("'","")
    return f1

if __name__ == "__main__":
    app.run(debug=True, port= 7329, host= "localhost")

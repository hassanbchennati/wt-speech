import openpyxl
import formulas

def AramcoSheet1 (csvpath): 
          
    '''
    Read Aramco cvs data: How much NGL we can produce in UGP if Khuff diverted to LP Letdown?  
    and take all function needed for the calulation to process it with the user input
    
    Attributes:
    -----------
        csvpath: path to the csv file 
    '''
    
    
    # read the excel sheet 
    # read the excel sheet 
    wb = openpyxl.load_workbook(csvpath)
    ws = wb['Input']

    # 
    sheet1function = {}
    sheet1inputs = {}
    function_coordinate = []
    funcResults = {}
    count = 0

    # read the function needed 
    for row in ws.iter_rows():
        flag = False
        count = 0 
        for cell in row:
            if ws[cell.coordinate].data_type == 's':
                name = ws[cell.coordinate].value 
                count += 1
            if ws[cell.coordinate].data_type == 'f':
#                 print(cell.coordinate)
#                 print(count)
                function = ws[cell.coordinate].value
                if count ==1 or count ==2 :
                    if (len(cell.coordinate)) < 4 : 
                        sheet1function[name] = function
                        function_coordinate.append(cell.coordinate)
                        count = 0 

    # read the input needed
    for row in ws.iter_rows():
        for cell in row:
            if ws[cell.coordinate].data_type == 's':
                name = ws[cell.coordinate].value 
            if ws[cell.coordinate].data_type == 'n':
                if ws[cell.coordinate].value != None: 
                    #print(cell.coordinate)
                    cell = cell.coordinate
                    sheet1inputs[name] = cell

    # take the input from the user 
    inputlist = []
    for x in sheet1inputs: 
        userinput= input(x)
        inputlist.append(userinput)

    #calculate the function 

    fpath, dir_output = csvpath, 'output'  
    xl_model = formulas.ExcelModel().loads(fpath).finish()


    for i in sheet1function.values():

        funarrgements = []
        func = formulas.Parser().ast(i)[1].compile()
        #print("function input neede",list(func.inputs))
        for cell in list(func.inputs):
            # if the function input is a none
             if ws[cell].value == None: 
                funarrgements.append(0)

             # if the function input is a user input 

             if cell in sheet1inputs.values(): 
                if cell == sheet1inputs["Enter Associated Gas Feed Rate"] : 
                    funarrgements.append(inputlist[0])
                elif cell == sheet1inputs["Enter Khuff to LP Letdown"] : 
                    funarrgements.append(inputlist[1])


            # if the function input is other function  
            # check if the function calculated before and take the result 
             if cell in function_coordinate: 
                for x in funcResults: 
                    if cell == x : 
                        funarrgements.append(funcResults[x])


               # print(ws[cell])
             elif ws[cell].data_type == 'f':
                #print ("function is ", ws[cell].value)
                func2 = formulas.Parser().ast(ws[cell].value)[1].compile()
                funarrgements2 = []
               # print("arr needed" , list(func2.inputs))
                for ar2 in list(func2.inputs):
                    if ar2 in function_coordinate:  
                        for x in funcResults: 

                            if ar2 == x : 
                                funarrgements2.append(funcResults[x])


                    elif ws[cell].value == None: 
                        funarrgements2.append(1)

                if len(funarrgements2) == 1: 
                    test= str (func2 (funarrgements2[0]))
                    funarrgements.append(test)


                elif len(funarrgements2) == 2: 
                    test= str (func2 (funarrgements2[0], funarrgements2[1]))
                    funarrgements.append(test)

                elif len(funarrgements2) == 3: 
                    test= str (func2 (funarrgements2[0], funarrgements2[1], funarrgements2[2]))
                    funarrgements.append(test)

                elif len(funarrgements2) == 4: 
                    test= str (func2 (funarrgements2[0], funarrgements2[1], funarrgements2[2], funarrgements2[3]))
                    funarrgements.append(test)





        # all parameter now found 

      #  print("count",count)
      #  print("funarrgements" , funarrgements)
        if len(funarrgements) == 1: 
            funcResults[function_coordinate[count]]= str (func (funarrgements[0]))
           # print(funcResults[function_coordinate[count]])
        elif len(funarrgements) == 2: 
            funcResults[function_coordinate[count]]= str (func (funarrgements[0], funarrgements[1]))

        elif len(funarrgements) == 3: 
            funcResults[function_coordinate[count]] =str (func (funarrgements[0], funarrgements[1],funarrgements[2] ))
           # print(funcResults[function_coordinate[count]])
        elif len(funarrgements) == 4: 
            funcResults[function_coordinate[count]] = str (func (funarrgements[0], funarrgements[1],funarrgements[2], funarrgements[3] ))
         #   print(funcResults[function_coordinate[count]])

        elif len(funarrgements) == 5: 
            funcResults[function_coordinate[count]] = str (func (funarrgements[0], funarrgements[1],funarrgements[2], funarrgements[3], funarrgements[4] ))
        #break 

        count +=1

    funcResultsFinal = {}
    count2 = 0 
    for i in sheet1function:
        funcResultsFinal[i] = funcResults[function_coordinate[count2]]
        count2 +=1 
        
        

    return(funcResultsFinal)



test=AramcoSheet1("How much NGL we can produce in UGP.xlsm")
text = 'TOTAL feed rate to LRUs is ' + str(test['TOTAL feed rate to LRUs']) + ","
text = text + 'Intercepted C2 Recovery' + test['Intercepted %C2 Recovery'] + ","
text = text + 'NGL Production ' + test['NGL Production '] + ","
text = text + 'Additional NGL following Khuff to LP Letdown' + test['Additional NGL following Khuff to LP Letdown'] + ","
text = text + '$ Value of the additional NGL' + test['$ Value of the additional NGL']

# print(test['TOTAL feed rate to LRUs'])
# print(test['Intercepted %C2 Recovery'])
# print(test['NGL Production '])
# print(test['Additional NGL following Khuff to LP Letdown'])
# print(test['$ Value of the additional NGL'])


print(text)